#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import sys
from dataclasses import dataclass
from datetime import date, datetime, timedelta, tzinfo, timezone
from decimal import Decimal
from pathlib import Path
from typing import Dict, List, Optional
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

try:
    from zoneinfo import ZoneInfo, ZoneInfoNotFoundError
except Exception:
    ZoneInfo = None
    class ZoneInfoNotFoundError(Exception):
        pass

NS = {"p": "urn:k461-dke-de:profile_generic-1", "e": "urn:k461-dke-de:extension-1"}
ROOT_CLOSE = "</ns1:object>"
OBIS_LABELS = {
    "0100010800ff": "1.8.0 Bezug (kWh)",
    "0100020800ff": "2.8.0 Einspeisung (kWh)",
}
OBIS_BEZUG = "0100010800ff"
OBIS_EINSPEISUNG = "0100020800ff"


class EuropeBerlinFallback(tzinfo):
    def _last_sunday(self, year: int, month: int) -> date:
        d = date(year + (month == 12), 1 if month == 12 else month + 1, 1) - timedelta(days=1)
        while d.weekday() != 6:
            d -= timedelta(days=1)
        return d

    def _is_dst_local(self, dt: datetime) -> bool:
        naive = dt.replace(tzinfo=None)
        start_day = self._last_sunday(dt.year, 3)
        end_day = self._last_sunday(dt.year, 10)
        start = datetime(dt.year, 3, start_day.day, 2, 0, 0)
        end = datetime(dt.year, 10, end_day.day, 3, 0, 0)
        return start <= naive < end

    def utcoffset(self, dt: Optional[datetime]) -> Optional[timedelta]:
        return timedelta(hours=2 if dt and self._is_dst_local(dt) else 1)

    def dst(self, dt: Optional[datetime]) -> Optional[timedelta]:
        return timedelta(hours=1 if dt and self._is_dst_local(dt) else 0)

    def tzname(self, dt: Optional[datetime]) -> str:
        return "CEST" if dt and self._is_dst_local(dt) else "CET"

    def fromutc(self, dt: datetime) -> datetime:
        if dt.tzinfo is not self:
            raise ValueError("fromutc: dt.tzinfo is not self")
        year = dt.year
        march = self._last_sunday(year, 3)
        october = self._last_sunday(year, 10)
        dst_start_utc = datetime(year, 3, march.day, 1, 0, 0)
        dst_end_utc = datetime(year, 10, october.day, 1, 0, 0)
        naive_utc = dt.replace(tzinfo=None)
        offset = timedelta(hours=2 if dst_start_utc <= naive_utc < dst_end_utc else 1)
        return (naive_utc + offset).replace(tzinfo=self)


def get_local_tz() -> tzinfo:
    if ZoneInfo is not None:
        try:
            return ZoneInfo("Europe/Berlin")
        except ZoneInfoNotFoundError:
            pass
        except Exception:
            pass
    return EuropeBerlinFallback()


LOCAL_TZ = get_local_tz()
UTC = timezone.utc


@dataclass
class Reading:
    capture_time_utc: datetime
    capture_time_local: datetime
    value_kwh: float


@dataclass
class DailyEndValue:
    day: date
    used_timestamp_local: datetime
    values: Dict[str, float]


@dataclass
class DailyTariffRow:
    day: date
    start_00_local: Optional[datetime]
    go_ende_05_local: Optional[datetime]
    ende_folgetag_00_local: Optional[datetime]
    bezug_00_kwh: Optional[float]
    bezug_05_kwh: Optional[float]
    bezug_next_00_kwh: Optional[float]
    go_kwh: Optional[float]
    standard_kwh: Optional[float]
    gesamt_kwh: Optional[float]


class ParseError(Exception):
    pass


def extract_embedded_xml(raw_bytes: bytes) -> str:
    start = raw_bytes.find(b"<?xml")
    if start < 0:
        raise ParseError("Kein XML-Start ('<?xml') in der Datei gefunden.")
    decoded = raw_bytes[start:].decode("utf-8", errors="ignore")
    end = decoded.rfind(ROOT_CLOSE)
    if end < 0:
        raise ParseError(f"XML-Ende ({ROOT_CLOSE}) nicht gefunden.")
    return decoded[: end + len(ROOT_CLOSE)]


def extract_capture_objects(root: ET.Element) -> Dict[str, str]:
    capture_objects = root.find("p:attributes/p:capture_objects", NS)
    if capture_objects is None:
        raise ParseError("capture_objects nicht gefunden.")
    mapping: Dict[str, str] = {}
    for obj in capture_objects.findall("p:capture_object", NS):
        obj_id = obj.attrib.get("id")
        logical_name = obj.findtext("e:logical_name", default="", namespaces=NS).strip()
        if obj_id and logical_name:
            mapping[obj_id] = logical_name.split(".")[0].lower()
    if not mapping:
        raise ParseError("Keine capture_objects / OBIS-Kennungen gefunden.")
    return mapping


def parse_signed_entries(column: ET.Element) -> List[Reading]:
    entries: List[Reading] = []
    for entry in column.findall("p:entry_gateway_signed", NS):
        capture_time_text = entry.findtext("e:capture_time", default="", namespaces=NS).strip()
        long64_text = entry.findtext("e:value/e:long64", default="", namespaces=NS).strip()
        scaler_text = entry.findtext("e:scaler", default="0", namespaces=NS).strip()
        if not capture_time_text or not long64_text:
            continue
        try:
            raw_value = Decimal(long64_text)
            scaler = int(scaler_text)
            capture_time_utc = datetime.fromisoformat(capture_time_text.replace("Z", "+00:00")).astimezone(UTC)
        except Exception as exc:
            raise ParseError(f"Fehler beim Parsen eines Eintrags: {exc}") from exc
        unit_text = entry.findtext("e:unit", default="", namespaces=NS).strip()
        unit = int(unit_text) if unit_text else None
        value_in_source_unit = raw_value * (Decimal(10) ** scaler)
        value_kwh = value_in_source_unit / Decimal(1000) if unit == 30 else value_in_source_unit
        entries.append(Reading(capture_time_utc, capture_time_utc.astimezone(LOCAL_TZ), float(value_kwh)))
    entries.sort(key=lambda x: x.capture_time_utc)
    return entries


def parse_file(input_path: Path) -> Dict[str, List[Reading]]:
    root = ET.fromstring(extract_embedded_xml(input_path.read_bytes()))
    capture_mapping = extract_capture_objects(root)
    simple_data = root.find("p:attributes/p:buffer/p:simple_data", NS)
    if simple_data is None:
        raise ParseError("simple_data im buffer nicht gefunden.")
    result: Dict[str, List[Reading]] = {}
    for column in simple_data.findall("p:column", NS):
        obis = capture_mapping.get(column.attrib.get("id", ""))
        if obis:
            result[obis] = parse_signed_entries(column)
    if not result:
        raise ParseError("Keine Messreihen in den Spalten gefunden.")
    return result


def build_time_index(readings: List[Reading], hour: int, minute: int = 0) -> Dict[date, Reading]:
    result: Dict[date, Reading] = {}
    for r in readings:
        lt = r.capture_time_local
        if lt.hour == hour and lt.minute == minute:
            current = result.get(lt.date())
            if current is None or lt < current.capture_time_local:
                result[lt.date()] = r
    return result


def build_daily_end_values(series_by_obis: Dict[str, List[Reading]]) -> List[DailyEndValue]:
    midnight_index: Dict[str, Dict[date, Reading]] = {}
    for obis, readings in series_by_obis.items():
        midnight_index[obis] = build_time_index(readings, 0, 0)
    available_followup_days = None
    for date_map in midnight_index.values():
        days = set(date_map.keys())
        available_followup_days = days if available_followup_days is None else (available_followup_days & days)
    if not available_followup_days:
        return []
    results: List[DailyEndValue] = []
    for followup_day in sorted(available_followup_days):
        target_day = followup_day - timedelta(days=1)
        values: Dict[str, float] = {}
        used_timestamp_local: Optional[datetime] = None
        for obis, date_map in midnight_index.items():
            reading = date_map[followup_day]
            values[obis] = reading.value_kwh
            if used_timestamp_local is None:
                used_timestamp_local = reading.capture_time_local
        if used_timestamp_local is not None:
            results.append(DailyEndValue(target_day, used_timestamp_local, values))
    return results


def diff_if_possible(a: Optional[float], b: Optional[float]) -> Optional[float]:
    if a is None or b is None:
        return None
    return b - a


def build_daily_tariff_rows(series_by_obis: Dict[str, List[Reading]]) -> List[DailyTariffRow]:
    bezug = series_by_obis.get(OBIS_BEZUG, [])
    bezug_00 = build_time_index(bezug, 0, 0)
    bezug_05 = build_time_index(bezug, 5, 0)

    all_days = set(bezug_00.keys()) | set(bezug_05.keys())
    if not all_days:
        return []

    min_day = min(all_days)
    max_day = max(all_days) - timedelta(days=1)
    if max_day < min_day:
        return []

    rows: List[DailyTariffRow] = []
    current = min_day
    while current <= max_day:
        next_day = current + timedelta(days=1)
        b00 = bezug_00.get(current)
        b05 = bezug_05.get(current)
        b24 = bezug_00.get(next_day)
        rows.append(
            DailyTariffRow(
                day=current,
                start_00_local=b00.capture_time_local if b00 else None,
                go_ende_05_local=b05.capture_time_local if b05 else None,
                ende_folgetag_00_local=b24.capture_time_local if b24 else None,
                bezug_00_kwh=b00.value_kwh if b00 else None,
                bezug_05_kwh=b05.value_kwh if b05 else None,
                bezug_next_00_kwh=b24.value_kwh if b24 else None,
                go_kwh=diff_if_possible(b00.value_kwh if b00 else None, b05.value_kwh if b05 else None),
                standard_kwh=diff_if_possible(b05.value_kwh if b05 else None, b24.value_kwh if b24 else None),
                gesamt_kwh=diff_if_possible(b00.value_kwh if b00 else None, b24.value_kwh if b24 else None),
            )
        )
        current += timedelta(days=1)

    return rows


def autosize_worksheet(ws) -> None:
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = max(len("" if c.value is None else str(c.value)) for c in column_cells)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)


def fmt_dt(dt: Optional[datetime]) -> Optional[str]:
    return dt.strftime("%Y-%m-%d %H:%M:%S") if dt else None


def apply_header_style(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_excel(output_path: Path, daily_rows: List[DailyEndValue], tariff_rows: List[DailyTariffRow]) -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Tagesendwerte"
    obis_order = [obis for obis in (OBIS_BEZUG, OBIS_EINSPEISUNG) if any(obis in row.values for row in daily_rows)]
    ws.append(["Datum", "verwendeter Zeitstempel lokal", *[OBIS_LABELS.get(obis, obis) for obis in obis_order]])
    apply_header_style(ws)
    for row in daily_rows:
        ws.append([
            row.day.isoformat(),
            row.used_timestamp_local.strftime("%Y-%m-%d %H:%M:%S"),
            *[row.values.get(obis) for obis in obis_order],
        ])
    for r in ws.iter_rows(min_row=2, min_col=3, max_col=2 + len(obis_order)):
        for cell in r:
            cell.number_format = "0.0000"
    autosize_worksheet(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    tariff = wb.create_sheet("Tarifzonen")
    tariff.append([
        "Datum",
        "Messzeit 00:00 lokal",
        "Messzeit 05:00 lokal",
        "Messzeit Folgetag 00:00 lokal",
        "Bezug 00:00 (kWh)",
        "Bezug 05:00 (kWh)",
        "Bezug Folgetag 00:00 (kWh)",
        "Go-Verbrauch 00-05 (kWh)",
        "Standard-Verbrauch 05-24 (kWh)",
        "Gesamtverbrauch 00-24 (kWh)",
    ])
    apply_header_style(tariff)
    for row in tariff_rows:
        tariff.append([
            row.day.isoformat(),
            fmt_dt(row.start_00_local),
            fmt_dt(row.go_ende_05_local),
            fmt_dt(row.ende_folgetag_00_local),
            row.bezug_00_kwh,
            row.bezug_05_kwh,
            row.bezug_next_00_kwh,
            row.go_kwh,
            row.standard_kwh,
            row.gesamt_kwh,
        ])
    for r in tariff.iter_rows(min_row=2, min_col=5, max_col=10):
        for cell in r:
            cell.number_format = "0.0000"
    autosize_worksheet(tariff)
    tariff.freeze_panes = "A2"
    tariff.auto_filter.ref = tariff.dimensions

    info = wb.create_sheet("Definition")
    info["A1"] = "Definitionen"
    info["A1"].font = Font(bold=True)
    info["A3"] = "Tagesendwert"
    info["A3"].font = Font(bold=True)
    info["A4"] = "Der Tagesendwert eines Tages D ist der erste vorhandene kumulative Zählerstand am Folgetag D+1 um 00:00 Uhr lokaler Zeit (Europe/Berlin)."
    info["A6"] = "Tarifzonen für Intelligent Octopus Go"
    info["A6"].font = Font(bold=True)
    info["A7"] = "Go-Zeit = 00:00 bis 05:00 lokaler Zeit des Kalendertags D."
    info["A8"] = "Standardzeit = 05:00 bis 00:00 lokaler Zeit des Folgetags D+1."
    info["A10"] = "Berechnung Bezug"
    info["A10"].font = Font(bold=True)
    info["A11"] = "Go-Verbrauch = Bezug(05:00) - Bezug(00:00)"
    info["A12"] = "Standard-Verbrauch = Bezug(Folgetag 00:00) - Bezug(05:00)"
    info["A13"] = "Gesamtverbrauch = Bezug(Folgetag 00:00) - Bezug(00:00)"
    info["A15"] = "Wichtiger Hinweis"
    info["A15"].font = Font(bold=True)
    info["A16"] = "Gesamtverbrauch wird direkt aus den beiden Tagesrand-Zählerständen berechnet. Er ist damit rechnerisch identisch zu Go-Verbrauch + Standard-Verbrauch, sofern alle drei Messpunkte vorhanden sind."
    info["A17"] = "Falls für einen Tag ein benötigter Messpunkt fehlt (00:00 oder 05:00), bleibt die berechnete Spalte leer."
    info.column_dimensions["A"].width = 140

    wb.save(output_path)


def write_csv(output_path: Path, daily_rows: List[DailyEndValue], tariff_rows: List[DailyTariffRow]) -> None:
    with output_path.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow([
            "Datum",
            "Tagesendwert Zeitstempel lokal",
            "1.8.0 Tagesendwert (kWh)",
            "2.8.0 Tagesendwert (kWh)",
            "Messzeit 00:00 lokal",
            "Messzeit 05:00 lokal",
            "Messzeit Folgetag 00:00 lokal",
            "Bezug 00:00 (kWh)",
            "Bezug 05:00 (kWh)",
            "Bezug Folgetag 00:00 (kWh)",
            "Go-Verbrauch 00-05 (kWh)",
            "Standard-Verbrauch 05-24 (kWh)",
            "Gesamtverbrauch 00-24 (kWh)",
        ])
        tariff_by_day = {row.day: row for row in tariff_rows}
        end_by_day = {row.day: row for row in daily_rows}
        all_days = sorted(set(tariff_by_day.keys()) | set(end_by_day.keys()))
        for day in all_days:
            end = end_by_day.get(day)
            tariff = tariff_by_day.get(day)
            writer.writerow([
                day.isoformat(),
                fmt_dt(end.used_timestamp_local) if end else "",
                f"{end.values.get(OBIS_BEZUG, 0.0):.4f}" if end and OBIS_BEZUG in end.values else "",
                f"{end.values.get(OBIS_EINSPEISUNG, 0.0):.4f}" if end and OBIS_EINSPEISUNG in end.values else "",
                fmt_dt(tariff.start_00_local) if tariff else "",
                fmt_dt(tariff.go_ende_05_local) if tariff else "",
                fmt_dt(tariff.ende_folgetag_00_local) if tariff else "",
                f"{tariff.bezug_00_kwh:.4f}" if tariff and tariff.bezug_00_kwh is not None else "",
                f"{tariff.bezug_05_kwh:.4f}" if tariff and tariff.bezug_05_kwh is not None else "",
                f"{tariff.bezug_next_00_kwh:.4f}" if tariff and tariff.bezug_next_00_kwh is not None else "",
                f"{tariff.go_kwh:.4f}" if tariff and tariff.go_kwh is not None else "",
                f"{tariff.standard_kwh:.4f}" if tariff and tariff.standard_kwh is not None else "",
                f"{tariff.gesamt_kwh:.4f}" if tariff and tariff.gesamt_kwh is not None else "",
            ])


def build_default_output_path(input_path: Path) -> Path:
    suffixes = "".join(input_path.suffixes)
    base = input_path.name[:-len(suffixes)] if suffixes else input_path.stem
    return input_path.with_name(f"{base}_tagesendwerte_mit_tarifzonen_bezug.xlsx")


def parse_args(argv: List[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extrahiert Tagesendwerte sowie Go-/Standard-Verbrauch (nur Bezug) aus einer SMGW HAN .sm_data.xml.cms Datei."
    )
    parser.add_argument("input_file", type=Path, help="Pfad zur .sm_data.xml.cms Datei")
    parser.add_argument("-o", "--output", type=Path, help="Pfad zur Ausgabe-XLSX-Datei")
    return parser.parse_args(argv)


def main(argv: List[str]) -> int:
    args = parse_args(argv)
    input_path: Path = args.input_file
    output_xlsx: Path = args.output or build_default_output_path(input_path)
    output_csv: Path = output_xlsx.with_suffix(".csv")

    if not input_path.exists():
        print(f"Fehler: Datei nicht gefunden: {input_path}", file=sys.stderr)
        return 2

    try:
        series_by_obis = parse_file(input_path)
        daily_rows = build_daily_end_values(series_by_obis)
        tariff_rows = build_daily_tariff_rows(series_by_obis)
        if not daily_rows and not tariff_rows:
            raise ParseError("Keine auswertbaren Tagesdaten gefunden.")
        write_excel(output_xlsx, daily_rows, tariff_rows)
        write_csv(output_csv, daily_rows, tariff_rows)
    except ParseError as exc:
        print(f"Fehler beim Auswerten der Datei: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"Unerwarteter Fehler: {exc}", file=sys.stderr)
        return 1

    print(f"Excel geschrieben: {output_xlsx}")
    print(f"CSV geschrieben:   {output_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
